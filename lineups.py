"""
NHL Lineups Scraper - Daily Faceoff -> Excel + JSON
Usage personnel uniquement. Un onglet par equipe.
Version 4 - production ready pour GitHub Actions
"""

import time
import random
import re
import json
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_URL = "https://www.dailyfaceoff.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

COLOR_HEADER_BG   = "1A3A5C"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_LINE_BG     = "D6E4F0"
COLOR_PP_BG       = "FFF2CC"
COLOR_DEF_BG      = "E2EFDA"
COLOR_GOALIE_BG   = "FCE4D6"

TEAMS = [
    ("anaheim-ducks",          "Anaheim Ducks"),
    ("boston-bruins",          "Boston Bruins"),
    ("buffalo-sabres",         "Buffalo Sabres"),
    ("calgary-flames",         "Calgary Flames"),
    ("carolina-hurricanes",    "Carolina Hurricanes"),
    ("chicago-blackhawks",     "Chicago Blackhawks"),
    ("colorado-avalanche",     "Colorado Avalanche"),
    ("columbus-blue-jackets",  "Columbus Blue Jackets"),
    ("dallas-stars",           "Dallas Stars"),
    ("detroit-red-wings",      "Detroit Red Wings"),
    ("edmonton-oilers",        "Edmonton Oilers"),
    ("florida-panthers",       "Florida Panthers"),
    ("los-angeles-kings",      "Los Angeles Kings"),
    ("minnesota-wild",         "Minnesota Wild"),
    ("montreal-canadiens",     "Montreal Canadiens"),
    ("nashville-predators",    "Nashville Predators"),
    ("new-jersey-devils",      "New Jersey Devils"),
    ("new-york-islanders",     "New York Islanders"),
    ("new-york-rangers",       "New York Rangers"),
    ("ottawa-senators",        "Ottawa Senators"),
    ("philadelphia-flyers",    "Philadelphia Flyers"),
    ("pittsburgh-penguins",    "Pittsburgh Penguins"),
    ("san-jose-sharks",        "San Jose Sharks"),
    ("seattle-kraken",         "Seattle Kraken"),
    ("st-louis-blues",         "St Louis Blues"),
    ("tampa-bay-lightning",    "Tampa Bay Lightning"),
    ("toronto-maple-leafs",    "Toronto Maple Leafs"),
    ("utah-mammoth",           "Utah Mammoth"),
    ("vancouver-canucks",      "Vancouver Canucks"),
    ("vegas-golden-knights",   "Vegas Golden Knights"),
    ("washington-capitals",    "Washington Capitals"),
    ("winnipeg-jets",          "Winnipeg Jets"),
]


def get_soup(url):
    time.sleep(random.uniform(2.0, 4.0))
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


def get_players_in_section(section_div):
    players = []
    for a in section_div.find_all("a", href=re.compile(r"/players/news/")):
        span = a.find("span")
        if span:
            name = span.get_text(strip=True)
            if name:
                players.append(name)
    return players


def find_section_by_title(soup, *keywords):
    for kw in keywords:
        for tag in soup.find_all(string=re.compile(kw, re.IGNORECASE)):
            container = tag.find_parent("div")
            for _ in range(8):
                if container is None:
                    break
                players = get_players_in_section(container)
                if players:
                    return players, container
                container = container.find_parent("div")
    return [], None


def scrape_team_lineup(team_slug):
    url = f"{BASE_URL}/teams/{team_slug}/line-combinations/"
    try:
        soup = get_soup(url)
    except Exception as e:
        print(f"  Erreur pour {team_slug}: {e}")
        return None

    data = {
        "forwards": {},
        "defence":  {},
        "pp_units": {},
        "goalies":  [],
    }

    title_patterns = {
        "forwards":  [r"Forward Lines", r"Forwards"],
        "defensive": [r"Defensive Pairings", r"Defense"],
        "pp1":       [r"1st Powerplay", r"Power Play 1", r"PP1"],
        "pp2":       [r"2nd Powerplay", r"Power Play 2", r"PP2"],
        "goalies":   [r"Goalies", r"Goalie"],
    }

    all_divs = soup.find_all("div")
    section_markers = []

    for div in all_divs:
        text = div.get_text(strip=True)
        for key, patterns in title_patterns.items():
            for pat in patterns:
                if re.match(pat, text, re.IGNORECASE) and len(text) < 50:
                    if len(div.find_all("div")) < 5:
                        section_markers.append((key, div, text))
                        break

    print(f"    Sections trouvees: {[(s[0], s[2]) for s in section_markers]}")

    sections = {}
    for key, marker_div, title in section_markers:
        container = marker_div.find_parent("div")
        for _ in range(6):
            if container is None:
                break
            players = get_players_in_section(container)
            if len(players) >= 2:
                sections[key] = players
                break
            container = container.find_parent("div")

    print(f"    Joueurs par section: { {k: len(v) for k,v in sections.items()} }")

    # Avants
    fwd = sections.get("forwards", [])
    if not fwd:
        all_names = []
        seen = set()
        for a in soup.find_all("a", href=re.compile(r"/players/news/")):
            sp = a.find("span")
            if sp:
                n = sp.get_text(strip=True)
                if n and n not in seen:
                    all_names.append(n)
                    seen.add(n)
        fwd = all_names[:12]

    for i in range(4):
        chunk = fwd[i*3:(i+1)*3]
        if len(chunk) == 3:
            data["forwards"][f"L{i+1}"] = {"LW": chunk[0], "C": chunk[1], "RW": chunk[2]}

    # Defenseurs
    defs = sections.get("defensive", [])
    for i in range(3):
        chunk = defs[i*2:(i+1)*2]
        if len(chunk) == 2:
            data["defence"][f"D{i+1}"] = {"LD": chunk[0], "RD": chunk[1]}

    # PP Units
    pp1 = sections.get("pp1", [])
    pp2 = sections.get("pp2", [])
    if not pp1:
        pp1, _ = find_section_by_title(soup, r"1st Powerplay", r"1st Power Play")
    if not pp2:
        pp2, _ = find_section_by_title(soup, r"2nd Powerplay", r"2nd Power Play")
    if pp1:
        data["pp_units"]["PP1"] = pp1[:5]
    if pp2:
        data["pp_units"]["PP2"] = pp2[:5]

    # Gardiens
    goalie_players = sections.get("goalies", [])
    if not goalie_players:
        goalie_players, _ = find_section_by_title(soup, r"Goalies", r"Goalie")
    data["goalies"] = goalie_players[:2]

    return data


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hcell(ws, row, col, text, bg=COLOR_HEADER_BG):
    c = ws.cell(row, col, text)
    c.font = Font(bold=True, color=COLOR_HEADER_FONT, name="Arial", size=11)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c


def dcell(ws, row, col, text, bg=None):
    c = ws.cell(row, col, text)
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = thin_border()
    return c


def write_team_sheet(wb, team_data, team_name):
    ws = wb.create_sheet(title=team_name[:31])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    row = 1

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"NHL {team_name.upper()} - Lineups")
    c.font = Font(bold=True, name="Arial", size=14, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"Mis a jour : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(italic=True, name="Arial", size=9, color="666666")
    c.alignment = Alignment(horizontal="right")
    row += 2

    # Avants
    for col, lbl in enumerate(["Ligne", "LW", "C", "RW", "", ""], 1):
        hcell(ws, row, col, lbl)
    ws.row_dimensions[row].height = 20
    row += 1

    if team_data and team_data["forwards"]:
        for key in ["L1", "L2", "L3", "L4"]:
            p = team_data["forwards"].get(key, {})
            c = ws.cell(row, 1, key)
            c.font = Font(bold=True, name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=COLOR_LINE_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            dcell(ws, row, 2, p.get("LW", "-"), COLOR_LINE_BG)
            dcell(ws, row, 3, p.get("C",  "-"), COLOR_LINE_BG)
            dcell(ws, row, 4, p.get("RW", "-"), COLOR_LINE_BG)
            dcell(ws, row, 5, "", COLOR_LINE_BG)
            dcell(ws, row, 6, "", COLOR_LINE_BG)
            row += 1
    else:
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1, "Donnees non disponibles").font = Font(italic=True, color="999999", name="Arial", size=10)
        row += 1
    row += 1

    # Defenseurs
    for col, lbl in enumerate(["Pairing", "LD", "RD", "", "", ""], 1):
        hcell(ws, row, col, lbl)
    ws.row_dimensions[row].height = 20
    row += 1

    if team_data and team_data["defence"]:
        for key in ["D1", "D2", "D3"]:
            p = team_data["defence"].get(key, {})
            c = ws.cell(row, 1, key)
            c.font = Font(bold=True, name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=COLOR_DEF_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            dcell(ws, row, 2, p.get("LD", "-"), COLOR_DEF_BG)
            dcell(ws, row, 3, p.get("RD", "-"), COLOR_DEF_BG)
            for col in [4, 5, 6]:
                dcell(ws, row, col, "", COLOR_DEF_BG)
            row += 1
    else:
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1, "Donnees non disponibles").font = Font(italic=True, color="999999", name="Arial", size=10)
        row += 1
    row += 1

    # Power Play
    for col, lbl in enumerate(["PP", "J1", "J2", "J3", "J4", "J5"], 1):
        hcell(ws, row, col, lbl)
    ws.row_dimensions[row].height = 20
    row += 1

    if team_data and team_data["pp_units"]:
        for key in ["PP1", "PP2"]:
            players = team_data["pp_units"].get(key, [])
            c = ws.cell(row, 1, key)
            c.font = Font(bold=True, name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=COLOR_PP_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            for ci, p in enumerate(players[:5], 2):
                dcell(ws, row, ci, p, COLOR_PP_BG)
            for ci in range(len(players) + 2, 8):
                dcell(ws, row, ci, "-", COLOR_PP_BG)
            row += 1
    else:
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1, "Donnees non disponibles").font = Font(italic=True, color="999999", name="Arial", size=10)
        row += 1
    row += 1

    # Gardiens
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "GARDIENS")
    c.font = Font(bold=True, name="Arial", size=11)
    c.fill = PatternFill("solid", start_color=COLOR_GOALIE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    goalies = team_data["goalies"] if team_data else []
    if goalies:
        for i, g in enumerate(goalies, 1):
            label = "Partant" if i == 1 else "Reserviste"
            dcell(ws, row, 1, label, COLOR_GOALIE_BG)
            ws.merge_cells(f"B{row}:F{row}")
            dcell(ws, row, 2, g, COLOR_GOALIE_BG)
            row += 1
    else:
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1, "N/A").font = Font(italic=True, color="999999", name="Arial", size=10)


def main():
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\nNHL Lineups Scraper v4 - {today}")
    print("=" * 50)

    wb = Workbook()
    wb.remove(wb.active)

    all_data = {"last_updated": today, "teams": {}}

    for slug, name in TEAMS:
        print(f"  Scraping : {name}...")
        team_data = scrape_team_lineup(slug)
        write_team_sheet(wb, team_data, name)
        all_data["teams"][slug] = {
            "name":     name,
            "forwards": team_data["forwards"] if team_data else {},
            "defence":  team_data["defence"]  if team_data else {},
            "pp_units": team_data["pp_units"] if team_data else {},
            "goalies":  team_data["goalies"]  if team_data else [],
        }

    wb.save("nhl_lineups.xlsx")
    print("  Excel cree : nhl_lineups.xlsx")

    with open("nhl_lineups.json", "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print("  JSON cree  : nhl_lineups.json")

    print(f"\nTermine! {len(TEAMS)} equipes mises a jour.")


if __name__ == "__main__":
    main()
